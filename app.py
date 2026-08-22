# 🌟 PREMIUM UI TRICK: Completely hide the GitHub Fork button and Cat Logo
import streamlit as st
from google import genai
from google.genai import types
import PIL.Image
import io
import os
import re
import json
import asyncio
import datetime
import pandas as pd
import edge_tts
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from st_copy_to_clipboard import st_copy_to_clipboard

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="NexGen AI Studio")
    
# 1. 页面基本配置
st.set_page_config(page_title="AI File & Receipt Assistant", layout="centered")

# ==========================================
# 🌐 多语言文本字典 (i18n Dictionary)
# ==========================================
LANG_DICT = {
    "🇨🇳 中文": {
        "app_title": "📄 AI 多功能文件与发票助手",
        "login_title": "🔐 AI 分析器用户通道",
        "tab_login": "🔑 用户登录",
        "tab_register": "📝 新用户注册",
        "lbl_user": "📞 手机号 / 用户名：",
        "lbl_pass": "🔑 密码：",
        "btn_login": "立即登录",
        "login_success": "✅ 登录成功！",
        "login_fail": "❌ 用户名或密码错误！",
        "reg_info": "💡 注册即可获得【每天 2 次免费 AI 深度分析】额度！",
        "reg_u_lbl": "📞 输入您的手机号或用户名：",
        "reg_p_lbl": "🔑 设置访问密码：",
        "reg_p2_lbl": "🔑 确认访问密码：",
        "btn_register": "提交注册并自动登录",
        "err_empty": "⚠️ 用户名和密码不能为空！",
        "err_mismatch": "⚠️ 两次输入的密码不一致！",
        "err_exists": "⚠️ 该账号已被注册！",
        "reg_success": "🎉 注册成功！已为您自动登录。",
        "quota_label": "今日剩余可用额度：",
        "quota_exhausted": "⚠️ 您今日的免费额度已用尽！请明日再来，或联系客服开通无限次通道。",
        "vip_card_title": "💎 升级 VIP 享受无限次极速分析：",
        "vip_card_desc": "限时特惠仅需 <b>RM 5.90 / 月</b>（每天 50 次深度分析 + 批量导出 Excel）。",
        "vip_btn_text": "📲 点击联系 WhatsApp 客服开通充值",
        "wa_msg_template": "你好！我想为账号【{}】充值额度 / 开通 VIP。",
        "phone_tip": "⚠️ 手机端温馨提示：为防止手机直接拍照导致网页刷新，建议您【先用手机相机拍好文件】，再点击下方按钮前往【相册】选取上传！",
        "upload_label": "📷 选择单张或多张文件/发票（支持单次多选）",
        "read_success": "已成功读取 {} 个文件！",
        "file_page": "📷 文件第 {} 页: {}",
        "file_type_label": "🔮 请选择文件类型：",
        "mode_receipt": "🧾 车辆/商业发票收据 (支持批量导出Excel)",
        "mode_medical": "🏥 健康检测报告",
        "mode_contract": "📄 商业合同与通用文件",
        "mode_custom": "✍️ 自由输入/其他全新文件",
        "prompt_label": "💬 您想对 AI 提问什么？（可直接修改或补充）",
        "btn_start_analysis": "🚀 开始 AI 深度分析",
        "analyzing": "AI 正在深度分析中，请稍候...",
        "res_header": "📊 AI 分析与提取结果",
        "edit_tip": "💡 **提示**：您可以直接**双击下方表格中的任何单元格**修改金额、店名或分类，下载的 Excel 将自动以修改后的数据为准！",
        "cat_board": "📈 **分类开销汇总看板**：",
        "btn_dl_excel": "📥 立即下载 Excel 记账汇总表 (.xlsx)",
        "excel_downloaded": "✅ Excel 报表已成功下载！",
        "excel_path_tip": "💡 文件已保存至您手机/电脑的【下载 (Downloads)】文件夹中。",
        "voice_title": "🎵 语音朗读报告：",
        "share_header": "📲 结果快捷分享通道",
        "share_radio_lbl": "📌 请选择您希望分享到 WhatsApp 的内容类型：",
        "share_opt_text": "📝 发送文字报告",
        "share_opt_voice": "🎵 发送语音报告 (MP3)",
        "wa_step1": "💡 步骤一：输入接收人电话并锁定：",
        "wa_phone_lbl": "📞 接收人电话 (选填，例如: 60123456789，留空则手动选择):",
        "wa_btn_lock": "🔒 锁定号码并生成通道",
        "wa_lock_ok": "✅ 号码 {} 锁定成功！",
        "wa_lock_empty": "✅ 已锁定为空号模式！",
        "wa_step2_txt": "📋 步骤二：点击下方复制文本报告：",
        "btn_copy": "📋 点击此处 ➡️ 一键复制 AI 分析文本",
        "copy_ok": "🎉 复制成功！请前往 WhatsApp 粘贴发送！",
        "wa_step2_voice": "💾 步骤二：点击下方下载语音文件：",
        "btn_dl_voice": "⬇️ 下载语音文件 (voice_report.mp3)",
        "voice_downloaded": "✅ 语音文件已下载至 Downloads 文件夹！",
        "voice_tip": "💡 提示：下载后打开 WhatsApp，在聊天框点击 📎 附件选择该 MP3 发送即可。",
        "wa_step3": "🟢 步骤三：点击前往 WhatsApp",
        "tts_voice": "zh-CN-YunxiNeural",
        "col_date": "日期",
        "col_merchant": "商家名称",
        "col_category": "分类",
        "col_invoice": "单据号码",
        "col_tax": "税额 (SST)",
        "col_total": "总金额 (Total)",
        "sheet1_name": "发票收据明细",
        "sheet2_name": "分类统计汇总",
        "sheet2_col1": "消费类别",
        "sheet2_col2": "分类汇总金额 (RM)",
        "total_row_label": "总计 (TOTAL)",
        "total_count_label": "共 {} 张单据",
    },
    "🇬🇧 English": {
        "app_title": "📄 AI Multi-Purpose Document & Receipt Assistant",
        "login_title": "🔐 AI Assistant User Portal",
        "tab_login": "🔑 Login",
        "tab_register": "📝 Register",
        "lbl_user": "📞 Phone Number / Username:",
        "lbl_pass": "🔑 Password:",
        "btn_login": "Login Now",
        "login_success": "✅ Login successful!",
        "login_fail": "❌ Invalid username or password!",
        "reg_info": "💡 Register now to get 2 FREE AI deep analyses daily!",
        "reg_u_lbl": "📞 Enter your phone number or username:",
        "reg_p_lbl": "🔑 Set password:",
        "reg_p2_lbl": "🔑 Confirm password:",
        "btn_register": "Submit & Login",
        "err_empty": "⚠️ Username and password cannot be empty!",
        "err_mismatch": "⚠️ Passwords do not match!",
        "err_exists": "⚠️ Username already registered!",
        "reg_success": "🎉 Registered successfully! Logged in automatically.",
        "quota_label": "Remaining daily quota:",
        "quota_exhausted": "⚠️ Your daily free quota is exhausted! Please return tomorrow or contact support.",
        "vip_card_title": "💎 Upgrade to VIP for Unlimited Deep Analysis:",
        "vip_card_desc": "Limited offer at only <b>RM 5.90 / month</b> (50 deep analyses/day + batch Excel export).",
        "vip_btn_text": "📲 Contact WhatsApp Support to Recharge",
        "wa_msg_template": "Hi! I would like to recharge / upgrade VIP for account [{}].",
        "phone_tip": "⚠️ Mobile Tip: To prevent camera auto-refresh, please take photos first and upload from your Photo Library!",
        "upload_label": "📷 Upload Documents / Receipts (Multiple files supported)",
        "read_success": "Successfully loaded {} files!",
        "file_page": "📷 File Page {}: {}",
        "file_type_label": "🔮 Select Document Type:",
        "mode_receipt": "🧾 Receipts & Invoices (Batch Excel Export)",
        "mode_medical": "🏥 Health Lab Report",
        "mode_contract": "📄 Business Contract & General Document",
        "mode_custom": "✍️ Custom Input / Other Documents",
        "prompt_label": "💬 What would you like to ask AI? (Editable)",
        "btn_start_analysis": "🚀 Start AI Deep Analysis",
        "analyzing": "AI is analyzing, please wait...",
        "res_header": "📊 Analysis & Extraction Results",
        "edit_tip": "💡 **Tip**: Double-click any table cell below to edit amounts or categories. The downloaded Excel will reflect your edits!",
        "cat_board": "📈 **Expense Summary by Category**:",
        "btn_dl_excel": "📥 Download Excel Expense Report (.xlsx)",
        "excel_downloaded": "✅ Excel report downloaded successfully!",
        "excel_path_tip": "💡 Saved to your Downloads folder.",
        "voice_title": "🎵 Voice Audio Report:",
        "share_header": "📲 Quick WhatsApp Share Channel",
        "share_radio_lbl": "📌 Select content type to share on WhatsApp:",
        "share_opt_text": "📝 Send Text Report",
        "share_opt_voice": "🎵 Send Voice Audio (MP3)",
        "wa_step1": "💡 Step 1: Enter & lock recipient phone number:",
        "wa_phone_lbl": "📞 Recipient Phone (Optional, e.g. 60123456789, leave blank for manual selection):",
        "wa_btn_lock": "🔒 Lock Number & Create Link",
        "wa_lock_ok": "✅ Phone {} locked successfully!",
        "wa_lock_empty": "✅ Blank number mode locked!",
        "wa_step2_txt": "📋 Step 2: Click below to copy text report:",
        "btn_copy": "📋 Click here ➡️ Copy AI Report Text",
        "copy_ok": "🎉 Copied! Please paste in WhatsApp!",
        "wa_step2_voice": "💾 Step 2: Download voice file below:",
        "btn_dl_voice": "⬇️ Download Voice File (voice_report.mp3)",
        "voice_downloaded": "✅ Voice file downloaded to Downloads!",
        "voice_tip": "💡 Tip: Open WhatsApp, click 📎 Attachment ➡️ Audio to send.",
        "wa_step3": "🟢 Step 3: Open WhatsApp",
        "tts_voice": "en-US-AndrewMultilingualNeural",
        "col_date": "Date",
        "col_merchant": "Merchant Name",
        "col_category": "Category",
        "col_invoice": "Receipt No.",
        "col_tax": "SST / Tax",
        "col_total": "Total Amount",
        "sheet1_name": "Receipt Details",
        "sheet2_name": "Category Summary",
        "sheet2_col1": "Expense Category",
        "sheet2_col2": "Total Amount (RM)",
        "total_row_label": "TOTAL",
        "total_count_label": "Total {} receipts",
    },
    "🇲🇾 Bahasa Melayu": {
        "app_title": "📄 Pembantu AI Dokumen & Resit Pelbagai Guna",
        "login_title": "🔐 Saluran Pengguna Pembantu AI",
        "tab_login": "🔑 Log Masuk",
        "tab_register": "📝 Daftar Baru",
        "lbl_user": "📞 No. Telefon / Nama Pengguna:",
        "lbl_pass": "🔑 Kata Laluan:",
        "btn_login": "Log Masuk Sekarang",
        "login_success": "✅ Log masuk berjaya!",
        "login_fail": "❌ Nama pengguna atau kata laluan tidak sah!",
        "reg_info": "💡 Daftar sekarang untuk dapat 2 kuota analisis AI percuma setiap hari!",
        "reg_u_lbl": "📞 Masukkan no telefon atau nama pengguna:",
        "reg_p_lbl": "🔑 Tetapkan kata laluan:",
        "reg_p2_lbl": "🔑 Sahkan kata laluan:",
        "btn_register": "Hantar & Log Masuk",
        "err_empty": "⚠️ Nama pengguna dan kata laluan tidak boleh kosong!",
        "err_mismatch": "⚠️ Kata laluan tidak sepadan!",
        "err_exists": "⚠️ Akaun telah didaftarkan!",
        "reg_success": "🎉 Pendaftaran berjaya! Telah log masuk secara automatik.",
        "quota_label": "Baki kuota hari ini:",
        "quota_exhausted": "⚠️ Kuota percuma anda hari ini telah habis! Sila kembali esok.",
        "vip_card_title": "💎 Naik Taraf ke VIP untuk Analisis Tanpa Had:",
        "vip_card_desc": "Tawaran terhad hanya <b>RM 5.90 / bulan</b> (50 analisis/hari + eksport Excel berkelompok).",
        "vip_btn_text": "📲 Hubungi WhatsApp Khidmat Pelanggan untuk Tambah Nilai",
        "wa_msg_template": "Salam! Saya ingin menambah kuota / menaik taraf VIP untuk akaun [{}].",
        "phone_tip": "⚠️ Tip Telefon: Untuk elak muat semula laman web, sila tangkap gambar dahulu dan muat naik dari Galeri!",
        "upload_label": "📷 Pilih Dokumen / Resit (Boleh pilih banyak fail)",
        "read_success": "Berjaya membaca {} fail!",
        "file_page": "📷 Fail Halaman {}: {}",
        "file_type_label": "🔮 Pilih Jenis Dokumen:",
        "mode_receipt": "🧾 Resit & Invois (Eksport Excel Berkelompok)",
        "mode_medical": "🏥 Laporan Kesihatan",
        "mode_contract": "📄 Kontrak Perniagaan & Dokumen Am",
        "mode_custom": "✍️ Input Bebas / Dokumen Lain",
        "prompt_label": "💬 Apa soalan anda untuk AI? (Boleh diubah)",
        "btn_start_analysis": "🚀 Mula Analisis AI",
        "analyzing": "AI sedang menganalisis, sila tunggu...",
        "res_header": "📊 Hasil Analisis & Pengekstrakan AI",
        "edit_tip": "💡 **Tip**: Klik dua kali pada sel jadual untuk membetulkan amaun atau kategori. Excel yang dimuat turun akan mengikut data terkini anda!",
        "cat_board": "📈 **Papan Ringkasan Perbelanjaan Mengikut Kategori**:",
        "btn_dl_excel": "📥 Muat Turun Laporan Excel (.xlsx)",
        "excel_downloaded": "✅ Laporan Excel berjaya dimuat turun!",
        "excel_path_tip": "💡 Fail disimpan dalam folder Muat Turun (Downloads).",
        "voice_title": "🎵 Laporan Suara Audio:",
        "share_header": "📲 Saluran Perkongsian WhatsApp Pantas",
        "share_radio_lbl": "📌 Pilih jenis kandungan untuk dihantar ke WhatsApp:",
        "share_opt_text": "📝 Hantar Laporan Teks",
        "share_opt_voice": "🎵 Hantar Audio Suara (MP3)",
        "wa_step1": "💡 Langkah 1: Masukkan & kunci no telefon penerima:",
        "wa_phone_lbl": "📞 No Telefon Penerima (Pilihan, cth: 60123456789, biarkan kosong jika pilih manual):",
        "wa_btn_lock": "🔒 Kunci Nombor & Bina Saluran",
        "wa_lock_ok": "✅ Nombor {} berjaya dikunci!",
        "wa_lock_empty": "✅ Mod nombor kosong dikunci!",
        "wa_step2_txt": "📋 Langkah 2: Klik di bawah untuk salin teks laporan:",
        "btn_copy": "📋 Klik Di Sini ➡️ Salin Teks Laporan AI",
        "copy_ok": "🎉 Berjaya disalin! Sila tampal ke WhatsApp!",
        "wa_step2_voice": "💾 Langkah 2: Muat turun audio suara di bawah:",
        "btn_dl_voice": "⬇️ Muat Turun Fail Audio (voice_report.mp3)",
        "voice_downloaded": "✅ Fail audio berjaya dimuat turun!",
        "voice_tip": "💡 Tip: Buka WhatsApp, klik 📎 Lampiran ➡️ Audio untuk hantar.",
        "wa_step3": "🟢 Langkah 3: Pergi ke WhatsApp",
        "tts_voice": "ms-MY-OsmanNeural",
        "col_date": "Tarikh",
        "col_merchant": "Nama Peniaga",
        "col_category": "Kategori",
        "col_invoice": "No. Resit",
        "col_tax": "Cukai / SST",
        "col_total": "Jumlah Amaun",
        "sheet1_name": "Butiran Resit",
        "sheet2_name": "Ringkasan Kategori",
        "sheet2_col1": "Kategori Perbelanjaan",
        "sheet2_col2": "Jumlah Amaun (RM)",
        "total_row_label": "JUMLAH (TOTAL)",
        "total_count_label": "Sebanyak {} resit",
    }
}

# 移动端样式优化
st.markdown(
    """
    <style>
    [data-testid="stHeader"], [data-testid="stAppDeployButton"], header, .stAppHeader, a[href*="github.com"] {
        display: none !important;
        visibility: hidden !important;
        height: 0px !important;
        padding: 0px !important;
        margin: 0px !important;
    }
    
    [data-testid="stAppViewBlockContainer"], .main .block-container, div.block-container {
        padding-top: 0.5rem !important;  
        padding-bottom: 1rem !important;
        max-width: 730px !important; 
        padding-left: 8px !important;   
        padding-right: 8px !important;
    }

    p, span, label, .stText, div {
        word-break: break-word !important;
        letter-spacing: 0.02rem !important;
    }
    
    div[data-testid="stTextInput"] input {
        font-size: 16px !important;     
    }
    </style>
    """,
    unsafe_allow_html=True
)

# 顶部语言选择下拉框
selected_lang = st.selectbox(
    "🌐 Language / 语言 / Bahasa",
    ["🇨🇳 中文", "🇬🇧 English", "🇲🇾 Bahasa Melayu"],
    index=0
)
T = LANG_DICT[selected_lang]


# 1. 精美的弹出介绍气泡
if "English" in selected_lang:
    with st.popover("👋Click to view the NexGen app introduction.", use_container_width=True):
        col_img, col_txt = st.columns([1, 7])
        # 👈 左边列：放图片
        with col_img:
            st.image(
                "logo01.png", 
                width=50
            )
        # 👉 右边列：放你想展示的文字和标题
        with col_txt:
            st.markdown("### What NexGen can do?")
 
        st.write("🤖 Multilingual Support")
        st.write("🤖 Automated Intelligent Text Analysis")
        st.write("🤖 Display format for text, files, and voice results")
        st.write("🤖 Supports direct sending via WhatsApp")

    #第二部分
    col_img, col_txt = st.columns([1, 7])
        # 👈 左边列：放图片
        with col_img:
            st.image(
                "logo01.png", 
                width=50
            )
        # 👉 右边列：放你想展示的文字和标题
        with col_txt:
            st.markdown("### Application steps : ")
 
        st.write("👉 1. Register and login")
        st.write("👉 2. Select/Snap photo(s)")
        st.write("👉 3. Select Document Type: ")
        st.write("   🧾 Receipts & Invoices (Batch Excel Export)")
        st.write("   🏥 Health Lab Report")
        st.write("   📄 Business Contract & General Document")
        st.write("   ✍️ Custom Input / Other Documents")




               

elif "Melayu" in selected_lang:
    with st.popover("👋Klik untuk melihat pengenalan aplikasi NexGen.", use_container_width=True):
        col_img, col_txt = st.columns([1, 7])
        # 👈 左边列：放图片
        with col_img:
            st.image(
                "logo01.png", 
                width=50
            )
        # 👉 右边列：放你想展示的文字和标题
        with col_txt:
            st.markdown("### Apakah NexGen boleh buat?")
        st.write("🤖 Sokongan berbilang bahasa")
        st.write("🤖 Analisis teks pintar secara automatik")
        st.write("🤖 Format paparan untuk hasil teks, dokumen dan pertuturan")
        st.write("🤖 Sokongan untuk penghantaran terus melalui WhatsApp")
else:
    with st.popover("👋点击查看 NexGen 应用介绍", use_container_width=True):
        col_img, col_txt = st.columns([1, 7])
        # 👈 左边列：放图片
        with col_img:
            st.image(
                "logo01.png", 
                width=50
            )
        # 👉 右边列：放你想展示的文字和标题
        with col_txt:
            st.markdown("### NexGen可以做什么?")
        st.write("🤖 多語言支援")
        st.write("🤖 自动化智能文本分析")
        st.write("🤖 文字、文件及語音結果的顯示格式")
        st.write("🤖 支援透過 WhatsApp 直接發送")    
# ================= 📊 Google Sheets 数据库连接层 =================
@st.cache_resource
def get_gspread_client():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds_dict = dict(st.secrets["gcp_service_account"])
    creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
    return gspread.authorize(creds)

def get_user_sheet():
    client = get_gspread_client()
    return client.open_by_url(st.secrets["GSHEET_URL"]).sheet1

def get_all_users():
    try:
        sheet = get_user_sheet()
        records = sheet.get_all_records(value_render_option="FORMATTED_VALUE")
        users_dict = {}
        for idx, r in enumerate(records, start=2):
            raw_u = str(r.get("username", "")).strip()
            raw_p = str(r.get("password", "")).strip()
            
            user_entry = {
                "row": idx,
                "password": raw_p,
                "daily_limit": int(r["daily_limit"]) if str(r.get("daily_limit", "")).isdigit() else 2,
                "used_today": int(r["used_today"]) if str(r.get("used_today", "")).isdigit() else 0,
                "last_date": str(r.get("last_date", "")).strip()
            }
            users_dict[raw_u] = user_entry
            if raw_u.startswith("0"):
                users_dict[raw_u.lstrip("0")] = user_entry
            else:
                users_dict[f"0{raw_u}"] = user_entry
        return users_dict
    except Exception:
        return {}

# ================= 🔒 第一步：用户登录与自主注册系统 =================
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False
    st.session_state["current_user"] = None

if not st.session_state["authenticated"]:
    #st.title(T["login_title"])
    st.markdown(f'<h1 style="font-size: 18px;">{T["login_title"]}</h1>', unsafe_allow_html=True)
    
    tab_login, tab_register = st.tabs([T["tab_login"], T["tab_register"]])
    
    with tab_login:
        login_u = st.text_input(T["lbl_user"], key="login_username")
        login_p = st.text_input(T["lbl_pass"], type="password", key="login_password")
        if st.button(T["btn_login"], type="primary", use_container_width=True, key="btn_login_submit"):
            with st.spinner("..."):
                users = get_all_users()
                clean_u = login_u.strip()
                clean_p = login_p.strip()
                
                matched_user = None
                if clean_u in users:
                    matched_user = users[clean_u]
                elif clean_u.lstrip("0") in users:
                    matched_user = users[clean_u.lstrip("0")]
                elif f"0{clean_u}" in users:
                    matched_user = users[f"0{clean_u}"]
                
                if matched_user and str(matched_user["password"]).strip() == clean_p:
                    st.session_state["authenticated"] = True
                    st.session_state["current_user"] = clean_u
                    st.success(T["login_success"])
                    st.rerun()
                else:
                    st.error(T["login_fail"])
                    
    with tab_register:
        st.info(T["reg_info"])
        reg_u = st.text_input(T["reg_u_lbl"], key="reg_username")
        reg_p = st.text_input(T["reg_p_lbl"], type="password", key="reg_password")
        reg_p2 = st.text_input(T["reg_p2_lbl"], type="password", key="reg_password2")
        
        if st.button(T["btn_register"], use_container_width=True, key="btn_register_submit"):
            clean_reg_u = reg_u.strip()
            clean_reg_p = reg_p.strip()
            if not clean_reg_u or not clean_reg_p:
                st.warning(T["err_empty"])
            elif clean_reg_p != reg_p2.strip():
                st.warning(T["err_mismatch"])
            else:
                with st.spinner("..."):
                    users = get_all_users()
                    if clean_reg_u in users:
                        st.warning(T["err_exists"])
                    else:
                        sheet = get_user_sheet()
                        today_str = str(datetime.date.today())
                        sheet.append_row([clean_reg_u, clean_reg_p, 2, 0, today_str], value_input_option="RAW")
                        st.session_state["authenticated"] = True
                        st.session_state["current_user"] = clean_reg_u
                        st.success(T["reg_success"])
                        st.rerun()
    st.stop()
# =============================================================

# ================= 🔑 第二步：智能 API Key 初始化 =================
clients = []
if "GEMINI_API_KEY" in st.secrets and st.secrets["GEMINI_API_KEY"].strip():
    clients.append(genai.Client(api_key=st.secrets["GEMINI_API_KEY"].strip()))
if "GEMINI_API_KEY_BACKUP" in st.secrets and st.secrets["GEMINI_API_KEY_BACKUP"].strip():
    clients.append(genai.Client(api_key=st.secrets["GEMINI_API_KEY_BACKUP"].strip()))
if "GEMINI_API_KEY_TRIPLE" in st.secrets and st.secrets["GEMINI_API_KEY_TRIPLE"].strip():
    clients.append(genai.Client(api_key=st.secrets["GEMINI_API_KEY_TRIPLE"].strip()))

if not clients:
    st.warning("⚠️ API Key Error")
    st.stop()

if "key_index" not in st.session_state:
    st.session_state["key_index"] = 0

def compress_image(image_bytes: bytes, max_dimension: int = 1600, quality: int = 82) -> bytes:
    try:
        img = PIL.Image.open(io.BytesIO(image_bytes))
        if hasattr(img, '_getexif'):
            img = PIL.ImageOps.exif_transpose(img) if hasattr(PIL, 'ImageOps') else img
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
        img.thumbnail((max_dimension, max_dimension), PIL.Image.Resampling.LANCZOS)
        output_io = io.BytesIO()
        img.save(output_io, format="JPEG", quality=quality, optimize=True)
        return output_io.getvalue()
    except Exception:
        return image_bytes

# 🌟 生成多语言双 Sheet 财务级 Excel（内部统一标准字段）
def create_formatted_excel(df: pd.DataFrame, t_dict: dict) -> bytes:
    df_export = df.copy()
    
    df_export["tax_amount"] = pd.to_numeric(df_export["tax_amount"], errors="coerce").fillna(0.0)
    df_export["total_amount"] = pd.to_numeric(df_export["total_amount"], errors="coerce").fillna(0.0)
    
    tax_total = df_export["tax_amount"].sum()
    amount_total = df_export["total_amount"].sum()
    
    category_summary = df_export.groupby("category")["total_amount"].sum().reset_index()
    category_summary.columns = [t_dict["sheet2_col1"], t_dict["sheet2_col2"]]
    
    rename_map = {
        "date": t_dict["col_date"],
        "merchant": t_dict["col_merchant"],
        "category": t_dict["col_category"],
        "invoice_no": t_dict["col_invoice"],
        "tax_amount": t_dict["col_tax"],
        "total_amount": t_dict["col_total"]
    }
    df_detail_named = df_export.rename(columns=rename_map)
    
    total_row = {
        t_dict["col_date"]: t_dict["total_row_label"],
        t_dict["col_merchant"]: t_dict["total_count_label"].format(len(df_export)),
        t_dict["col_category"]: "-",
        t_dict["col_invoice"]: "-",
        t_dict["col_tax"]: tax_total,
        t_dict["col_total"]: amount_total
    }
    df_detail_final = pd.concat([df_detail_named, pd.DataFrame([total_row])], ignore_index=True)
    
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df_detail_final.to_excel(writer, index=False, sheet_name=t_dict["sheet1_name"])
        category_summary.to_excel(writer, index=False, sheet_name=t_dict["sheet2_name"])
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_font = Font(name="Calibri", size=11, bold=True, color="000000")
        regular_font = Font(name="Calibri", size=10)
        
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
        )
        double_bottom_border = Border(
            left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000')
        )

        ws1 = writer.sheets[t_dict["sheet1_name"]]
        for col in range(1, ws1.max_column + 1):
            cell = ws1.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in range(2, ws1.max_row + 1):
            is_total = (row == ws1.max_row)
            for col in range(1, ws1.max_column + 1):
                cell = ws1.cell(row=row, column=col)
                cell.font = total_font if is_total else regular_font
                cell.border = double_bottom_border if is_total else thin_border
                if is_total:
                    cell.fill = total_fill

                if col in (5, 6):
                    cell.number_format = '#,##0.00;(#,##0.00);0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col in (1, 3, 4):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        for col in ws1.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws1.column_dimensions[col_letter].width = max(max_len + 8, 16)

        ws2 = writer.sheets[t_dict["sheet2_name"]]
        for col in range(1, ws2.max_column + 1):
            cell = ws2.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in range(2, ws2.max_row + 1):
            for col in range(1, ws2.max_column + 1):
                cell = ws2.cell(row=row, column=col)
                cell.font = regular_font
                cell.border = thin_border
                if col == 2:
                    cell.number_format = '#,##0.00;(#,##0.00);0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws2.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws2.column_dimensions[col_letter].width = max(max_len + 8, 22)

    return excel_buffer.getvalue()

# ================= 📄 第三步：App 核心业务功能 =================
#st.title(T["app_title"])
st.markdown(f'<h1 style="font-size: 18px;">{T["app_title"]}</h1>', unsafe_allow_html=True)

users = get_all_users()
current_user = st.session_state["current_user"]
user_data = users.get(current_user, None)

today_str = str(datetime.date.today())
remaining_quota = 2

if user_data:
    if user_data["last_date"] != today_str:
        try:
            sheet = get_user_sheet()
            sheet.update_cell(user_data["row"], 4, 0)
            sheet.update_cell(user_data["row"], 5, today_str)
            user_data["used_today"] = 0
            user_data["last_date"] = today_str
        except Exception:
            pass

    remaining_quota = max(0, user_data["daily_limit"] - user_data["used_today"])
    st.caption(f"👤 `{current_user}` ｜ {T['quota_label']} **{remaining_quota} / {user_data['daily_limit']}**")
else:
    st.caption(f"👤 `{current_user}` ｜ {T['quota_label']} **{remaining_quota}**")

st.warning(T["phone_tip"])

uploaded_files = st.file_uploader(
    T["upload_label"], 
    type=["jpg", "jpeg", "png", "pdf"], 
    accept_multiple_files=True
)

ai_contents = []

if uploaded_files:
    st.success(T["read_success"].format(len(uploaded_files)))
    for i, uploaded_file in enumerate(uploaded_files):
        file_type = uploaded_file.name.split(".")[-1].lower()
        file_bytes = uploaded_file.read()
        
        if file_type == "pdf":
            ai_contents.append(types.Part.from_bytes(data=file_bytes, mime_type="application/pdf"))
            st.info(f"📁 {uploaded_file.name}")
        else:
            compressed_bytes = compress_image(file_bytes)
            ai_contents.append(types.Part.from_bytes(data=compressed_bytes, mime_type="image/jpeg"))
            try:
                preview_img = PIL.Image.open(io.BytesIO(compressed_bytes))
                st.image(preview_img, width=200, caption=T["file_page"].format(i+1, uploaded_file.name))
            except Exception:
                st.info(f"📷 {uploaded_file.name}")

if ai_contents:
    file_mode = st.selectbox(
        T["file_type_label"], 
        [T["mode_receipt"], T["mode_medical"], T["mode_contract"], T["mode_custom"]],
        key=f"file_mode_select_{selected_lang}"
    )
    
    # 动态匹配提示词
    if "English" in selected_lang:
        lang_instruction = "Please respond in English."
        if file_mode == T["mode_medical"]:
            default_prompt = (
                f"{lang_instruction}\n"
                "Health Lab Report:\n"
                "Please analyze the documents and reply strictly in 【English】:\n"
                "1. Provide personalized, practical dietary advice and daily health monitoring reminders based on my current results.\n"
            )
        elif file_mode == T["mode_receipt"]:
            default_prompt = (
                f"{lang_instruction}\n"
                "Please analyze the uploaded receipt(s)/invoice(s).\n"
                "[Task 1]: Provide a clear human-readable summary (merchant name, items, total amount).\n"
                "[Task 2 - CRITICAL]: At the very end, output a valid JSON array wrapped in ```json and ``` with exact keys:\n"
                "[\n"
                "  {\n"
                "    \"date\": \"YYYY-MM-DD\",\n"
                "    \"merchant\": \"Merchant Name\",\n"
                "    \"category\": \"Category (Petrol/Dining/Repair/Office/Others)\",\n"
                "    \"invoice_no\": \"Receipt No\",\n"
                "    \"tax_amount\": 0.00,\n"
                "    \"total_amount\": 0.00\n"
                "  }\n"
                "]"
            )
        elif file_mode == T["mode_contract"]:
            default_prompt = (
                f"{lang_instruction}\n"
                "Extract core information from the document:\n"
                "1. Main subject and key terms/clauses?\n"
                "2. Important dates & signing parties?\n"
                "3. Key responsibilities and potential risks?"
            )
        else:
            default_prompt = f"{lang_instruction}\nPlease analyze the uploaded file and summarize core points:\n1."

    elif "Melayu" in selected_lang:
        lang_instruction = "Sila berikan jawapan dalam Bahasa Melayu."
        if file_mode == T["mode_medical"]:
            default_prompt = (
                f"{lang_instruction}\n"
                "Laporan Kesihatan:\n"
                "Sila analisis dokumen-dokumen tersebut dan berikan maklum balas secara eksklusif dalam 【Bahasa Melayu】:\n"
                "1. Berikan nasihat pemakanan yang praktikal dan diperibadikan, serta peringatan pemantauan kesihatan harian berdasarkan keputusan semasa saya.\n"
            )
        elif file_mode == T["mode_receipt"]:
            default_prompt = (
                f"{lang_instruction}\n"
                "Sila analisa resit/invois yang dimuat naik.\n"
                "[Tugasan 1]: Berikan ringkasan teks yang jelas (nama peniaga, butiran, jumlah amaun).\n"
                "[Tugasan 2 - SANGAT PENTING]: Di bahagian paling akhir, berikan tatasusunan JSON yang sah dalam ```json dan ``` dengan kunci berikut:\n"
                "[\n"
                "  {\n"
                "    \"date\": \"YYYY-MM-DD\",\n"
                "    \"merchant\": \"Nama Peniaga\",\n"
                "    \"category\": \"Kategori (Minyak/Makanan/Pembaikan/Pejabat/Lain-lain)\",\n"
                "    \"invoice_no\": \"No Resit\",\n"
                "    \"tax_amount\": 0.00,\n"
                "    \"total_amount\": 0.00\n"
                "  }\n"
                "]"
            )
        elif file_mode == T["mode_contract"]:
            default_prompt = (
                f"{lang_instruction}\n"
                "Ekstrak maklumat utama dokumen:\n"
                "1. Tajuk utama dan syarat teras?\n"
                "2. Tarikh penting & pihak yang terlibat?\n"
                "3. Tanggungjawab utama & risiko penting?"
            )
        else:
            default_prompt = f"{lang_instruction}\nSila analisa dokumen ini dan ringkaskan isi utama:\n1."

    else:
        lang_instruction = "请使用简体中文回答。"
        if file_mode == T["mode_medical"]:
            default_prompt = (
                f"{lang_instruction}\n"
                "化验单关键信息提取任务：\n"
                "1. 提取肌酐（Creatinine）、尿素（Urea）、血红蛋白（Hb）核心数值，用Markdown表格列出并附参考值。\n"
                "2. 评估当前指标状态与健康注意事项。"
            )
            default_prompt = (
                f"{lang_instruction}\n"
                "健康检测报告:\n"
                "请分析这些文件，并严格使用【中文】回复。:\n"
                "1. 根據我目前的檢查結果，提供個人化、實用的飲食建議及日常健康監測提醒。\n"
            )
        elif file_mode == T["mode_receipt"]:
            default_prompt = (
                f"{lang_instruction}\n"
                "请分析我上传的发票/收据单据（可能包含单张或多张）。\n"
                "【任务 1】：输出清晰的人类易读文字总结（包括商户、明细、金额总计）。\n"
                "【任务 2 极为重要】：在回答的最末尾，输出一个用 ```json 与 ``` 包裹的标准 JSON 数组，包含所有单据的结构化数据，字段必须严格对应如下英文键名：\n"
                "[\n"
                "  {\n"
                "    \"date\": \"YYYY-MM-DD\",\n"
                "    \"merchant\": \"Merchant Name\",\n"
                "    \"category\": \"Category (e.g., Petrol / Dining / Repair / Office / Miscellaneous)\",\n"
                "    \"invoice_no\": \"Receipt No\",\n"
                "    \"tax_amount\": 0.00,\n"
                "    \"total_amount\": 0.00\n"
                "  }\n"
                "]"
            )
        elif file_mode == T["mode_contract"]:
            default_prompt = (
                f"{lang_instruction}\n"
                "提取整理文件核心信息：\n"
                "1. 核心条款与主题？\n"
                "2. 关键日期与主体？\n"
                "3. 核心责任与风险点？"
            )
        else:
            default_prompt = f"{lang_instruction}\n请根据我上传的文件，分析核心信息并总结要点：\n1."

    user_prompt = st.text_area(
        T["prompt_label"], 
        value=default_prompt,
        height=160,
        key=f"prompt_input_{selected_lang}_{file_mode}"
    )

    # 🌟 额度与充值逻辑
    ADMIN_PHONE = "60122382546" 
    wa_recharge_msg = T["wa_msg_template"].format(current_user)
    wa_recharge_url = f"https://wa.me/{ADMIN_PHONE}?text={wa_recharge_msg}"

    is_quota_empty = (remaining_quota <= 0)

    if is_quota_empty:
        st.error(T["quota_exhausted"])
        st.markdown(
            f"""
            <div style="background-color: #f8d7da; border: 1px solid #f5c6cb; border-radius: 8px; padding: 15px; margin-bottom: 15px;">
                <p style="color: #721c24; margin: 0 0 10px 0; font-weight: bold;">{T['vip_card_title']}</p>
                <p style="color: #721c24; margin: 0 0 12px 0; font-size: 14px;">{T['vip_card_desc']}</p>
                <a href="{wa_recharge_url}" target="_blank" style="
                    display: block;
                    width: 100%;
                    text-align: center;
                    background-color: #25D366;
                    color: white;
                    padding: 12px 0px;
                    font-size: 15px;
                    font-weight: bold;
                    text-decoration: none;
                    border-radius: 6px;
                ">{T['vip_btn_text']}</a>
            </div>
            """,
            unsafe_allow_html=True
        )

    # 唯一 key 绑定的开始分析按钮
    if st.button(T["btn_start_analysis"], type="primary", use_container_width=True, disabled=is_quota_empty, key="btn_do_analysis"):
        if is_quota_empty:
            st.stop()

        with st.spinner(T["analyzing"]):
            final_inputs = [*ai_contents, user_prompt]
            success = False
            
            for attempt in range(len(clients)):
                current_client = clients[(st.session_state["key_index"] + attempt) % len(clients)]
                try:
                    response = current_client.models.generate_content(
                        model='gemini-3.6-flash',
                        contents=final_inputs
                    )
                    raw_text = response.text
                    st.session_state["analysis_result"] = raw_text
                    
                    st.session_state["invoice_raw_df"] = None
                    json_match = re.search(r"```json\s*([\s\S]*?)\s*```", raw_text)
                    if json_match:
                        try:
                            json_data = json.loads(json_match.group(1).strip())
                            if isinstance(json_data, list) and len(json_data) > 0:
                                df = pd.DataFrame(json_data)
                                for expected_col in ["date", "merchant", "category", "invoice_no", "tax_amount", "total_amount"]:
                                    if expected_col not in df.columns:
                                        df[expected_col] = 0.0 if "amount" in expected_col else "-"
                                
                                df["tax_amount"] = pd.to_numeric(df["tax_amount"], errors="coerce").fillna(0.0)
                                df["total_amount"] = pd.to_numeric(df["total_amount"], errors="coerce").fillna(0.0)
                                st.session_state["invoice_raw_df"] = df
                        except Exception:
                            pass
                    
                    # 语音生成
                    try:
                        clean_voice_text = re.sub(r"```json[\s\S]*?```", "", raw_text)
                        clean_voice_text = clean_voice_text.replace("*", "").replace("#", "").replace("`", "").strip()
                        
                        target_voice = T["tts_voice"]
                        async def generate_voice_data(text_to_read: str) -> bytes:
                            communicator = edge_tts.Communicate(text_to_read, target_voice)
                            audio_stream = b""
                            async for chunk in communicator.stream():
                                if chunk["type"] == "audio":
                                    audio_stream += chunk["data"]
                            return audio_stream
                        
                        audio_data = asyncio.run(generate_voice_data(clean_voice_text))
                        if audio_data:
                            st.session_state["audio_bytes"] = audio_data
                    except Exception:
                        pass
                    
                    if user_data and "row" in user_data:
                        try:
                            sheet = get_user_sheet()
                            sheet.update_cell(user_data["row"], 4, user_data["used_today"] + 1)
                        except Exception:
                            pass
                    
                    st.session_state["clear_clipboard_trigger"] = True
                    success = True
                    st.session_state["key_index"] = (st.session_state["key_index"] + attempt) % len(clients)
                    st.rerun()
                    break
                    
                except Exception as e:
                    err_msg = str(e)
                    if "429" in err_msg or "RESOURCE_EXHAUSTED" in err_msg:
                        if attempt < len(clients) - 1:
                            continue
                        else:
                            st.error(f"⚠️ API Key Quota Limit: {err_msg}")
                    elif "503" in err_msg or "UNAVAILABLE" in err_msg:
                        st.error("⚠️ AI server busy, please try again.")
                    else:
                        st.error(f"Error: {err_msg}")
                        break

# ================= 🟢 第四步：展示结果与下载 Excel =================
if "analysis_result" in st.session_state and st.session_state["analysis_result"]:
    if st.session_state.get("clear_clipboard_trigger", False):
        st.markdown(
            """
            <script>
            if (navigator.clipboard && navigator.clipboard.writeText) {
                navigator.clipboard.writeText("");
            }
            </script>
            """,
            unsafe_allow_html=True
        )
        st.session_state["clear_clipboard_trigger"] = False

    st.divider()
    st.subheader(T["res_header"])
    
    if st.session_state.get("invoice_raw_df") is not None:
        df = st.session_state["invoice_raw_df"]
        st.info(T["edit_tip"])
        
        col_configs = {
            "date": st.column_config.TextColumn(T["col_date"]),
            "merchant": st.column_config.TextColumn(T["col_merchant"]),
            "category": st.column_config.TextColumn(T["col_category"]),
            "invoice_no": st.column_config.TextColumn(T["col_invoice"]),
            "tax_amount": st.column_config.NumberColumn(T["col_tax"], format="RM %.2f"),
            "total_amount": st.column_config.NumberColumn(T["col_total"], format="RM %.2f"),
        }
        
        edited_df = st.data_editor(
            df,
            use_container_width=True,
            num_rows="dynamic",
            column_config=col_configs,
            key="invoice_editor_safe"
        )
        
        edited_df["total_amount"] = pd.to_numeric(edited_df["total_amount"], errors="coerce").fillna(0.0)
        edited_df["tax_amount"] = pd.to_numeric(edited_df["tax_amount"], errors="coerce").fillna(0.0)
        
        cat_group = edited_df.groupby("category")["total_amount"].sum().reset_index()
        
        st.write(T["cat_board"])
        cols = st.columns(len(cat_group) if len(cat_group) > 0 else 1)
        for idx, row_cat in cat_group.iterrows():
            with cols[idx % len(cols)]:
                st.metric(label=f"🏷️ {row_cat['category']}", value=f"RM {row_cat['total_amount']:.2f}")

        excel_data = create_formatted_excel(edited_df, T)
        
        def notify_excel_download():
            st.toast(T["excel_downloaded"], icon="📥")

        st.download_button(
            label=T["btn_dl_excel"],
            data=excel_data,
            file_name=f"Expenses_Report_{datetime.date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            on_click=notify_excel_download,
            use_container_width=True,
            key="btn_download_excel"
        )
        st.caption(T["excel_path_tip"])
        st.write("")

    if "audio_bytes" in st.session_state and st.session_state["audio_bytes"]:
        st.write(T["voice_title"])
        st.audio(st.session_state["audio_bytes"], format="audio/mp3")
        st.write("")
    
    display_text = re.sub(r"```json[\s\S]*?```", "", st.session_state["analysis_result"]).strip()
    st.markdown(display_text)
    
    st.divider()
    st.subheader(T["share_header"])
    
    share_type = st.radio(
        T["share_radio_lbl"],
        options=[T["share_opt_text"], T["share_opt_voice"]],
        horizontal=True,
        key="share_type_radio"
    )
    
    with st.form("whatsapp_form", clear_on_submit=False):
        st.info(T["wa_step1"])
        target_phone = st.text_input(T["wa_phone_lbl"], value="")
        lock_button = st.form_submit_button(T["wa_btn_lock"], use_container_width=True)
        
        if lock_button:
            clean_phone = target_phone.strip().replace("+", "").replace(" ", "").replace("\t", "").replace("\n", "")
            if clean_phone:
                st.session_state["wa_url"] = f"https://wa.me/{clean_phone}"
                st.success(T["wa_lock_ok"].format(clean_phone))
            else:
                st.session_state["wa_url"] = "https://wa.me/"
                st.success(T["wa_lock_empty"])

    st.write("")

    if share_type == T["share_opt_text"]:
        st.write(T["wa_step2_txt"])
        st_copy_to_clipboard(
            display_text, 
            before_copy_label=T["btn_copy"], 
            after_copy_label=T["copy_ok"]
        )
        
    elif share_type == T["share_opt_voice"]:
        if "audio_bytes" in st.session_state and st.session_state["audio_bytes"]:
            st.write(T["wa_step2_voice"])
            
            def notify_voice_download():
                st.toast(T["voice_downloaded"], icon="🎵")

            st.download_button(
                label=T["btn_dl_voice"],
                data=st.session_state["audio_bytes"],
                file_name="voice_report.mp3",
                mime="audio/mp3",
                on_click=notify_voice_download,
                use_container_width=True,
                key="btn_download_voice"
            )
            st.caption(T["voice_tip"])
        else:
            st.warning("⚠️ No audio data.")

    if "wa_url" in st.session_state and st.session_state["wa_url"]:
        whatsapp_btn_html = f"""
        <a href="{st.session_state['wa_url']}" target="_blank" style="
            display: block; 
            width: 100%; 
            text-align: center; 
            background-color: #25D366; 
            color: white; 
            padding: 14px 0px; 
            font-size: 16px; 
            font-weight: bold; 
            text-decoration: none; 
            border-radius: 8px; 
            margin-top: 20px; 
            box-shadow: 0px 4px 10px rgba(37,211,102,0.3);
        ">{T['wa_step3']}</a>
        """
        st.markdown(whatsapp_btn_html, unsafe_allow_html=True)
